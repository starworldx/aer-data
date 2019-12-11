# The code in this file has dependencies.
# The code will only work if all dependencies are met.
# Otherwise code is for illustration only.

# This program parses AER daily drilling report text file and uploads into a database for further analysis.
# (In this case, MS Excel).

#! python3

from os import listdir
from os.path import isfile, join
from openpyxl import Workbook, load_workbook
import warnings
warnings.filterwarnings("ignore")

FILE = "book.xlsx"
WB = load_workbook(FILE)

class LicenceCancel:
     Date = None
     WellName = None
     LicenceNumber = None
     UID = None
     amendments = None
     def printLicence(self):
         print ("Date"," - ",self.Date)
         print ("Well Name"," - ",self.WellName)
         print ("Licence"," - ",self.LicenceNumber)
         print ("UID"," - ",self.UID)

     def writeLicenceToFile(self):
        wb = WB
        ws = wb['Cancelled']
        rows = ws.get_highest_row()+1
        cols = ws.get_highest_column()

        ws.cell(row=rows, column=1).value = self.Date
        ws.cell(row=rows, column=2).value = self.WellName
        ws.cell(row=rows, column=3).value = self.LicenceNumber
        ws.cell(row=rows, column=4).value = self.UID
        wb.save(FILE)

class LicenceAmend:
     Date = None
     WellName = None
     LicenceNumber = None
     UID = None
     amendments = None
     def printLicence(self):
         print ("Date"," - ",self.Date)
         print ("Well Name"," - ",self.WellName)
         print ("Licence"," - ",self.LicenceNumber)
         print ("UID"," - ",self.UID)
         for key in self.amendments:
              print (key," - ",self.amendments[key])

     def writeLicenceToFile(self):
        wb = WB
        ws = wb['Amended']
        rows = ws.get_highest_row()+1
        cols = ws.get_highest_column()

        ws.cell(row=rows, column=1).value = self.Date
        ws.cell(row=rows, column=2).value = self.WellName
        ws.cell(row=rows, column=3).value = self.Date
        ws.cell(row=rows, column=4).value = self.UID
        ws.cell(row=rows, column=5).value = self.LicenceNumber
        ##     	Well Type	Projected Depth	Ground Elevation	NUMBER OF DWELLINGS IN TH
        for key in self.amendments:
             if 'UWI' == key:
                  ws.cell(row=rows, column=6).value = self.amendments[key]

             elif 'WELL NAME' == key:
                  ws.cell(row=rows, column=7).value = self.amendments[key]

             elif 'Target Substance' == key:
                  ws.cell(row=rows, column=8).value = self.amendments[key]

             elif 'TARGET SUBSTANCE(1)' == key:
                  ws.cell(row=rows, column=9).value = self.amendments[key]

             elif 'TARGET SUBSTANCE(2)' == key:
                  ws.cell(row=rows, column=10).value = self.amendments[key]

             elif 'MAXIMUM CALCULATED EPZ' == key:
                  ws.cell(row=rows, column=11).value = self.amendments[key]

             elif 'DRILL CUTTINGS' == key:
                  temp = self.amendments[key].split(" TO ")
                  ws.cell(row=rows, column=12).value = temp[0]
                  ws.cell(row=rows, column=13).value = temp[1]

             elif 'REGULATION SECTION' == key:
                  ws.cell(row=rows, column=14).value = self.amendments[key]

             elif 'SURFACE CO-ORDINATES' == key:
                  temp = self.amendments[key].split("W")
                  if len(temp)==1:
                       temp = self.amendments[key].split("E")
                       temp[1] = "E"+temp[1]
                  else:
                       temp[1] = "W"+temp[1]
                  ws.cell(row=rows, column=15).value = temp[0]
                  ws.cell(row=rows, column=16).value = temp[1]

             elif 'WELL TYPE' == key:
                  ws.cell(row=rows, column=17).value = self.amendments[key]

             elif 'Projected Depth Ground Elevation' == key:
                  ws.cell(row=rows, column=18).value = self.amendments[key]

             elif 'NUMBER OF DWELLINGS IN TH' == key:
                  ws.cell(row=rows, column=19).value = self.amendments[key]

             else:
                  inserted = False
                  for row in ws.rows:
                       for i in range(0,len(row)):
                            if row[i].value == key:
                                 ws.cell(row=rows, column=i).value = self.amendments[key]
                                 inserted = True
                       break
                  if not inserted:
                       ws.cell(row=0, column=cols).value = key
                       ws.cell(row=rows, column=cols).value = self.amendments[key]

        wb.save(FILE)

class LicenceNew:
     Date = None
     WellName = None
     LicenceNumber = None
     MineralRights = None
     GroundElevation = None
     UniqueIdentifier = None
     SurfaceCoordinates = None
     BoardFieldCentre = None
     ProjectedDepth = None
     LaheeClassification = None
     Field = None
     TerminatingZone = None
     DrillingOperation = None
     WellPurpose = None
     WellType = None
     Substance = None
     Licensee= None
     SurfaceLocation= None

     def printLicence(self):
         print ("Date"," - ",self.Date)
         print ("Well Name"," - ",self.WellName)
         print ("Licence"," - ",self.LicenceNumber)
         print ("MineralRights"," - ",self.MineralRights)
         print ("GroundElevation"," - ",self.GroundElevation)
         print ("UniqueIdentifier"," - ",self.UniqueIdentifier)
         print ("SurfaceCoordinates"," - ",self.SurfaceCoordinates)
         print ("BoardFieldCentre"," - ",self.BoardFieldCentre)
         print ("ProjectedDepth"," - ",self.ProjectedDepth)
         print ("LaheeClassification"," - ",self.LaheeClassification)
         print ("Field"," - ",self.Field)
         print ("TerminatingZone"," - ",self.TerminatingZone)
         print ("DrillingOperation"," - ",self.DrillingOperation)
         print ("WellPurpose"," - ",self.WellPurpose)
         print ("WellType"," - ",self.WellType)
         print ("Substance"," - ",self.Substance)
         print ("Licensee"," - ",self.Licensee)
         print ("SurfaceLocation"," - ",self.SurfaceLocation)

     def writeLicenceToFile(self):
        wb = WB
        ws = wb['New Licenses']
        rows = ws.get_highest_row()+1
        cols = ws.get_highest_column()

        ws.cell(row=rows, column=1).value = self.Date
        ws.cell(row=rows, column=2).value = self.WellName
        ws.cell(row=rows, column=3).value = self.LicenceNumber
        ws.cell(row=rows, column=4).value = self.MineralRights
        ws.cell(row=rows, column=5).value = self.GroundElevation
        ws.cell(row=rows, column=6).value = self.UniqueIdentifier
        temp = self.SurfaceCoordinates.split("W")
        if len(temp)==1:
             temp = self.SurfaceCoordinates.split("E")
             temp[1] = "E"+temp[1]
        else:
             temp[1] = "W"+temp[1]

        ws.cell(row=rows, column=7).value = temp[0].strip()
        ws.cell(row=rows, column=8).value = temp[1].strip()
        ws.cell(row=rows, column=9).value = self.BoardFieldCentre
        ws.cell(row=rows, column=10).value = self.ProjectedDepth
        ws.cell(row=rows, column=11).value = self.LaheeClassification
        ws.cell(row=rows, column=12).value = self.Field
        ws.cell(row=rows, column=13).value = self.TerminatingZone
        ws.cell(row=rows, column=14).value = self.DrillingOperation
        ws.cell(row=rows, column=15).value = self.WellPurpose
        ws.cell(row=rows, column=16).value = self.WellType
        ws.cell(row=rows, column=17).value = self.Substance
        ws.cell(row=rows, column=18).value = self.Licensee
        ws.cell(row=rows, column=19).value = self.SurfaceLocation
        wb.save(FILE)

class SpudParser:
    ##            lic.printLicence()
    def getFilesInDir(self,dir):
        files = [f for f in listdir(dir) if isfile(join(dir, f))]
        return files
    def parseFile(self,file):
        newLicences = []
        amendLicences = []
        cancelLicences = []
        lines = tuple(open(file, 'r'))
        DATE =  lines[6].strip().replace("DATE: ","")
        lines = lines[17:]
        LL = 0;
        while LL+6 < len(lines) and "-----" not in lines[LL+1]:
            lic = LicenceNew()
            licenceLines = lines[LL:LL+5]
            lic.Date = DATE
            lic.WellName =  licenceLines[0][:41].strip()
            lic.LicenceNumber =  licenceLines[0][41:51].strip()
            lic.MineralRights =  licenceLines[0][51:72].strip()
            lic.GroundElevation =  licenceLines[0][72:].strip()

            lic.UniqueIdentifier =  licenceLines[1][:27].strip()
            lic.SurfaceCoordinates =  licenceLines[1][27:51].strip()
            lic.BoardFieldCentre =  licenceLines[1][51:72].strip()
            lic.ProjectedDepth =  licenceLines[1][72:].strip().replace("M","")

            lic.LaheeClassification =  licenceLines[2][:41].strip()
            lic.Field =  licenceLines[2][41:51].strip()
            lic.TerminatingZone =  licenceLines[2][72:].strip()

            lic.DrillingOperation =  licenceLines[3][:41].strip()
            lic.WellPurpose =  licenceLines[3][41:51].strip()
            lic.WellType =  licenceLines[3][51:72].strip()
            lic.Substance =  licenceLines[3][72:].strip()

            lic.Licensee =  licenceLines[4][:41].strip()
            lic.SurfaceLocation =  licenceLines[4][72:].strip()
            newLicences.append(lic)
            LL = LL + 6
            if "AMENDMENTS" in lines[LL]:
                 LL = LL+6
                 while "----" not in lines[LL+1]:
                      TL = LL+2
                      while 1:
                           if len(lines[TL]) > 1:
                                if len(lines[TL][0:15].strip()) > 0:
                                     break
                           TL = TL + 1

                      licenceLines = lines[LL:TL-1]
                      lic = LicenceAmend()
                      lic.Date = DATE
                      lic.WellName = licenceLines[0][:42].strip()
                      lic.LicenceNumber = licenceLines[0][42:53].strip()
                      lic.UID = licenceLines[1][:53].strip()
                      i=0
                      lic.amendments = {}
                      while i < len(licenceLines):
                           lic.amendments[licenceLines[i][53:].strip().replace(":", "")]=licenceLines[i+1][53:].strip()
                           i=i+3


                      amendLicences.append(lic)
                      LL = TL

            if "CANCELLED" in lines[LL]:
                 LL = LL+6
                 while "----" not in lines[LL+1]:
                      lic = LicenceCancel()
                      lic.Date = DATE
                      lic.WellName = lines[LL][:42].strip()
                      lic.LicenceNumber = lines[LL][42:53].strip()
                      lic.UID = lines[LL+1][:53].strip()
                      cancelLicences.append(lic)
                      LL = LL+3
        licences = {}
        licences['new'] = newLicences
        licences['amend'] = amendLicences
        licences['cancel'] = cancelLicences

        return licences

if __name__ == "__main__":

    PATH = ".\Daily File Downld"

    parser = SpudParser()
    files = parser.getFilesInDir(PATH)
    for f in files:
        print ("Saving File: ",f)
        licences = parser.parseFile(PATH+"\\"+f)
        for key in licences:
             print (key," = ",len(licences[key]))
        for lic in licences['new']:
            lic.writeLicenceToFile()
        for lic in licences['amend']:
            lic.writeLicenceToFile()
        for lic in licences['cancel']:
            lic.writeLicenceToFile()
    print ("done")
