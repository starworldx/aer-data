#! python3

from os import listdir
from os.path import isfile, join
from openpyxl import Workbook
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

class Contractor:
    NAME = None
    BAID = None

class Spud:
     WellId = None
     WellName = None
     Licence = None
     CONTRACTOR = None
     RigName = None
     ActivityDate = None
     FieldCenter = None
     BAID = None
     LICENSEE = None
     NPTD = None
     ActivityType= None

     def printSpud(self):
         print ("Well Id"," - ",self.WellId)
         print ("Well Name"," - ",self.WellName)
         print ("Licence"," - ",self.Licence)
         print ("CONTRACTOR BAID"," - ",self.CONTRACTOR.BAID)
         print ("CONTRACTOR NAME"," - ",self.CONTRACTOR.NAME)
         print ("Rig Name"," - ",self.RigName)
         print ("Activity Date"," - ",self.ActivityDate)
         print ("Field Center"," - ",self.FieldCenter)
         print ("BA ID"," - ",self.BAID)
         print ("LICENSEE"," - ",self.LICENSEE)
         print ("NEW PROJECTED TOTAL DEPTH"," - ",self.NPTD)
         print ("Activity Type"," - ",self.ActivityType)

     def writeSpudsToFile(self,file):
        wb = load_workbook(file)
        ws = wb['Spud Data']
        rows = ws.get_highest_row()+1
        cols = ws.get_highest_column()

        ws.cell(row=rows, column=1).value = self.WellId
        ws.cell(row=rows, column=2).value = self.WellName
        ws.cell(row=rows, column=3).value = self.Licence
        ws.cell(row=rows, column=4).value = self.CONTRACTOR.BAID
        ws.cell(row=rows, column=5).value = self.CONTRACTOR.NAME
        ws.cell(row=rows, column=6).value = self.RigName
        ws.cell(row=rows, column=7).value = self.ActivityDate
        ws.cell(row=rows, column=8).value = self.FieldCenter
        ws.cell(row=rows, column=9).value = self.BAID
        ws.cell(row=rows, column=10).value = self.LICENSEE
        ws.cell(row=rows, column=11).value = self.NPTD
        ws.cell(row=rows, column=12).value = self.ActivityType
        wb.save(file)
        pass

class SpudParser:
    def getFilesInDir(self,dir):
        files = [f for f in listdir(dir) if isfile(join(dir, f))]
        return files
    def parseFile(self,file):
        spuds = []
        lines = tuple(open(file, 'r'))
        for line in lines:
            if line[0]=='-':
                temp = line.split()
                for i in range(0,len(temp)):
                    temp[i] = len(temp[i])+1
                    if i>0:
                        temp[i] = temp[i]+temp[i-1]

            sp = Spud()
            sp.CONTRACTOR = Contractor()
            if " AM " in line or " PM " in line:
                sp.WellId = line[0:temp[0]]
                sp.WellName = line[temp[0]:temp[1]].strip()
                sp.Licence = line[temp[1]:temp[2]].strip()
                sp.CONTRACTOR.BAID = line[temp[2]:temp[3]].strip()
                sp.CONTRACTOR.NAME = line[temp[3]:temp[4]].strip()
                sp.RigName = line[temp[4]:temp[5]].strip()
                sp.ActivityDate = line[temp[5]:temp[6]].strip()
                sp.FieldCenter = line[temp[6]:temp[7]].strip()
                sp.BAID = line[temp[7]:temp[8]].strip()
                sp.LICENSEE = line[temp[8]:temp[9]].strip()
                sp.NPTD = line[temp[9]:temp[10]].strip()
                sp.ActivityType = line[temp[10]:temp[11]].strip()
                spuds.append(sp)
        return spuds


if __name__ == "__main__":

    PATH = "files\Spud"

    parser = SpudParser()
    files = parser.getFilesInDir(PATH)
    for f in files:
        print ("Saving File: ",f)
        spuds = parser.parseFile(PATH+"\\"+f)
        for spud in spuds:
            spud.writeSpudsToFile('book.xlsx')
        print ("Total Spuds Saved: %d" % len(spuds))
    print ("done")
